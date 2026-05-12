"""
Trading service - order management, position tracking, and sandbox execution.
"""
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import and_, desc, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.order import Order, OrderSide, OrderStatus, OrderType, Trade
from app.models.position import Position as PositionModel


# ---------------------------------------------------------------------------
# Order management
# ---------------------------------------------------------------------------

async def create_order(
    db: AsyncSession,
    user_id: int,
    order_data: Dict[str, Any],
) -> Order:
    """
    Create a new order with built-in risk and validation checks.

    In sandbox mode (controlled by *order_data.get("sandbox")*) the order
    is simulated rather than executed against a live exchange.

    Args:
        db: Database session.
        user_id: ID of the user placing the order.
        order_data: Dictionary with keys:
            symbol, side, type, quantity, price (limit orders),
            stop_price (stop orders), strategy_id, sandbox (optional).

    Returns:
        The created Order object.

    Raises:
        ValueError: If validation fails.
    """
    symbol = order_data["symbol"]
    side = OrderSide(order_data["side"])
    order_type = OrderType(order_data.get("type", OrderType.LIMIT.value))
    quantity = float(order_data["quantity"])
    price = float(order_data["price"]) if order_data.get("price") else None
    stop_price = float(order_data["stop_price"]) if order_data.get("stop_price") else None
    sandbox = order_data.get("sandbox", True)

    # -- Validation --
    # Price is required for limit orders
    if order_type == OrderType.LIMIT and price is None:
        raise ValueError("Price is required for limit orders")

    if quantity <= 0:
        raise ValueError("Quantity must be greater than 0")

    # Build order record
    order = Order(
        user_id=user_id,
        strategy_id=order_data.get("strategy_id"),
        symbol=symbol,
        side=side,
        type=order_type,
        status=OrderStatus.PENDING,
        price=price,
        stop_price=stop_price,
        quantity=quantity,
        filled_quantity=0.0,
        fee=0.0,
        source=order_data.get("source", "manual"),
        remark=order_data.get("remark"),
    )

    if sandbox:
        # Sandbox: simulate immediate execution
        _simulate_sandbox_execution(order)
    else:
        # 实盘模式：通过交易所适配器下单
        await _execute_via_adapter(order, db)

    db.add(order)
    await db.flush()
    await db.refresh(order)

    # Record trade if filled (sandbox or real)
    if order.status == OrderStatus.FILLED and order.avg_price:
        trade = Trade(
            order_id=order.id,
            symbol=order.symbol,
            side=order.side,
            price=order.avg_price,
            quantity=order.filled_quantity,
            fee=order.fee or 0,
            trade_time=datetime.now(timezone.utc),
        )
        db.add(trade)
        await db.flush()

    return order


async def cancel_order(
    db: AsyncSession,
    user_id: int,
    order_id: int,
) -> Optional[Order]:
    """
    Cancel a pending order.

    Only orders in PENDING or PARTIAL status can be cancelled.
    Users can only cancel their own orders.

    Args:
        db: Database session.
        user_id: ID of the user (for ownership check).
        order_id: ID of the order to cancel.

    Returns:
        The cancelled Order object if found and cancellable, None otherwise.
    """
    result = await db.execute(
        select(Order).where(
            and_(Order.id == order_id, Order.user_id == user_id)
        )
    )
    order = result.scalar_one_or_none()
    if order is None:
        return None

    if order.status not in (OrderStatus.PENDING, OrderStatus.PARTIAL):
        raise ValueError(
            f"Cannot cancel order in status '{order.status.value}'"
        )

    order.status = OrderStatus.CANCELED
    await db.flush()
    await db.refresh(order)
    return order


async def get_orders(
    db: AsyncSession,
    user_id: int,
    status: Optional[OrderStatus] = None,
    symbol: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
) -> List[Order]:
    """
    Query orders for a user with optional status filter and pagination.

    Args:
        db: Database session.
        user_id: ID of the user.
        status: Filter by order status (optional).
        skip: Number of records to skip.
        limit: Maximum number of records to return.

    Returns:
        A list of Order objects, ordered by creation time (newest first).
    """
    query = select(Order).where(Order.user_id == user_id)

    if status is not None:
        query = query.where(Order.status == status)
    if symbol is not None:
        query = query.where(Order.symbol == symbol)

    query = query.order_by(desc(Order.created_at)).offset(skip).limit(limit)
    result = await db.execute(query)
    return list(result.scalars().all())


# ---------------------------------------------------------------------------
# Position management
# ---------------------------------------------------------------------------

async def get_positions(
    db: AsyncSession,
    user_id: int,
    symbol: Optional[str] = None,
) -> List[PositionModel]:
    """
    Get all current positions for a user, with real-time price refresh.

    Args:
        db: Database session.
        user_id: ID of the user.

    Returns:
        A list of Position objects with up-to-date prices.
    """
    query = select(PositionModel).where(PositionModel.user_id == user_id)
    if symbol:
        query = query.where(PositionModel.symbol == symbol)
    query = query.order_by(PositionModel.symbol)
    result = await db.execute(query)
    positions = list(result.scalars().all())

    if positions:
        await _batch_refresh_positions(db, positions)

    return positions


async def get_position(
    db: AsyncSession,
    user_id: int,
    symbol: str,
) -> Optional[PositionModel]:
    """
    Get a single position for a user and symbol.

    Args:
        db: Database session.
        user_id: ID of the user.
        symbol: Trading symbol.

    Returns:
        The Position object if found, None otherwise.
    """
    result = await db.execute(
        select(PositionModel).where(
            and_(
                PositionModel.user_id == user_id,
                PositionModel.symbol == symbol,
            )
        )
    )
    position = result.scalar_one_or_none()
    if position:
        await _batch_refresh_positions(db, [position])
    return position


async def update_position(
    db: AsyncSession,
    user_id: int,
    symbol: str,
    position_data: Dict[str, Any],
) -> Optional[PositionModel]:
    """
    Update a position's fields (used internally after trade execution).

    If no position exists for the user+symbol, a new one is created.

    Args:
        db: Database session.
        user_id: ID of the user.
        symbol: Trading symbol.
        position_data: Dictionary of fields to update
            (quantity, available_quantity, cost_price, etc.).

    Returns:
        The updated or created Position object.
    """
    result = await db.execute(
        select(PositionModel).where(
            and_(
                PositionModel.user_id == user_id,
                PositionModel.symbol == symbol,
            )
        )
    )
    position = result.scalar_one_or_none()

    if position is None:
        position = PositionModel(
            user_id=user_id,
            symbol=symbol,
            quantity=0,
            available_quantity=0,
            frozen_quantity=0,
            cost_price=0,
            current_price=0,
        )
        db.add(position)

    for key, value in position_data.items():
        if hasattr(position, key):
            setattr(position, key, value)

    await db.flush()
    await db.refresh(position)
    return position


# ---------------------------------------------------------------------------
# Trade history
# ---------------------------------------------------------------------------

async def get_trades(
    db: AsyncSession,
    user_id: int,
    symbol: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
) -> List[Trade]:
    """
    Get trade history for a user.

    Args:
        db: Database session.
        user_id: ID of the user.
        symbol: Filter by trading symbol (optional).
        skip: Number of records to skip.
        limit: Maximum number of records to return.

    Returns:
        A list of Trade objects, newest first.
    """
    # Trades are linked to orders, which belong to users
    query = (
        select(Trade)
        .join(Order, Trade.order_id == Order.id)
        .where(Order.user_id == user_id)
    )
    if symbol:
        query = query.where(Trade.symbol == symbol)
    query = (
        query.order_by(desc(Trade.trade_time))
        .offset(skip)
        .limit(limit)
    )
    result = await db.execute(query)
    return list(result.scalars().all())


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

async def _execute_via_adapter(order: Order, db: AsyncSession) -> None:
    """通过交易所适配器执行真实下单"""
    import logging
    from app.core.config import settings
    from app.utils.exchange_adapter import ExchangeAdapterFactory

    logger = logging.getLogger(__name__)
    try:
        adapter = ExchangeAdapterFactory.create(settings.ORDER_EXECUTION_MODE)
        result = await adapter.create_order(
            symbol=order.symbol,
            side=order.side.value,
            order_type=order.type.value,
            quantity=order.quantity,
            price=order.price,
        )
        order.order_id_exchange = result.get("order_id") or result.get("entrust_no", "")
        order.avg_price = result.get("avg_price") or order.price or 0
        order.filled_quantity = result.get("filled_quantity", order.quantity)
        order.fee = result.get("fee", 0)
        # 实盘订单状态由后续定时任务同步（easytrader 返回不代表最终成交）
        status = result.get("status", "")
        if status in ("filled", "全部成交"):
            order.status = OrderStatus.FILLED
        elif status in ("pending", "已报", "部成"):
            order.status = OrderStatus.PARTIAL
        else:
            order.status = OrderStatus.PENDING
    except Exception as e:
        logger.error("实盘下单失败: %s", e)
        order.status = OrderStatus.REJECTED
        order.remark = (order.remark or "") + f" [拒绝原因: {e}]"


def _simulate_sandbox_execution(order: Order) -> None:
    """沙盒模式模拟成交。使用最新 ticker 价格作为成交价，无行情数据则拒绝订单。"""
    order.status = OrderStatus.FILLED
    order.filled_quantity = order.quantity

    if order.price is None or order.price <= 0:
        # Try to use the latest ticker price for realistic simulation
        try:
            from app.core.database import SyncSessionLocal
            from app.models.market_data import Ticker as TickerModel
            from sqlalchemy import select as sync_select

            sdb = SyncSessionLocal()
            try:
                result = sdb.execute(
                    sync_select(TickerModel).where(TickerModel.symbol == order.symbol)
                )
                ticker = result.scalar_one_or_none()
                if ticker and ticker.last_price and ticker.last_price > 0:
                    order.avg_price = round(float(ticker.last_price), 2)
                else:
                    raise ValueError(f"No ticker price available for {order.symbol}")
            finally:
                sdb.close()
        except Exception:
            import logging
            logging.getLogger(__name__).warning(
                "沙盒成交失败: 无法获取 %s 的行情价格，订单已拒绝（无ticker数据）",
                order.symbol,
            )
            order.status = OrderStatus.REJECTED
            order.remark = (order.remark or "") + " [拒绝: 模拟环境下无法获取标的行情]"
            return
    else:
        order.avg_price = order.price

    order.fee = round(order.avg_price * order.quantity * 0.001, 4)


async def check_risk_before_trade(db: AsyncSession, user_id: int, symbol: str, order_data: dict) -> dict:
    """交易前风控检查（代理到 risk 服务）"""
    from app.services.risk import check_risk_rules
    return await check_risk_rules(db, user_id, symbol, order_data)


# ---------------------------------------------------------------------------
# Manual position entry & Excel import
# ---------------------------------------------------------------------------

async def create_or_update_position(
    db: AsyncSession,
    user_id: int,
    symbol: str,
    quantity: float,
    cost_price: float,
    leverage: float = 1.0,
    position: PositionModel | None = None,
) -> PositionModel:
    """手动录入或更新持仓。存在则覆盖 quantity/cost_price，不存在则新建。

    Pass *position* to skip the existence lookup (callers that have already fetched it).
    """
    if position is None:
        result = await db.execute(
            select(PositionModel).where(
                and_(
                    PositionModel.user_id == user_id,
                    PositionModel.symbol == symbol,
                )
            )
        )
        position = result.scalar_one_or_none()

    market_value = round(quantity * cost_price * leverage, 2)

    if position is None:
        position = PositionModel(
            user_id=user_id,
            symbol=symbol,
            quantity=quantity,
            available_quantity=quantity,
            frozen_quantity=0,
            cost_price=cost_price,
            current_price=cost_price,
            market_value=market_value,
            pnl=0,
            pnl_ratio=0,
            day_pnl=0,
            day_pnl_ratio=0,
            margin=round(market_value / leverage, 2) if leverage > 0 else market_value,
            leverage=leverage,
        )
        db.add(position)
    else:
        position.quantity = quantity
        position.available_quantity = quantity
        position.cost_price = cost_price
        position.market_value = market_value
        position.margin = round(market_value / leverage, 2) if leverage > 0 else market_value
        position.leverage = leverage
        position.pnl = 0
        position.pnl_ratio = 0
        position.day_pnl = 0
        position.day_pnl_ratio = 0

    # 尝试用最新行情更新 current_price 和盈亏
    await _refresh_position_price(db, position)

    await db.flush()
    await db.refresh(position)
    return position


async def _batch_refresh_positions(db: AsyncSession, positions: list) -> None:
    """Batch-update all positions with latest ticker prices in a single round-trip."""
    if not positions:
        return

    from app.models.market_data import Ticker as TickerModel

    # Build OR query: (symbol == pos.symbol) OR (symbol LIKE pos.symbol.%)
    conditions = []
    for p in positions:
        conditions.append(TickerModel.symbol == p.symbol)
        conditions.append(TickerModel.symbol.like(p.symbol + ".%"))

    ticker_result = await db.execute(
        select(TickerModel).where(or_(*conditions))
    )
    all_tickers = list(ticker_result.scalars().all())

    # Build lookup: position symbol → best ticker
    ticker_map: dict[str, Any] = {}
    for t in all_tickers:
        short = t.symbol.rsplit(".", 1)[0] if "." in t.symbol else t.symbol
        # First match wins (exact match takes priority since it comes first in OR)
        if short not in ticker_map:
            ticker_map[short] = t

    for pos in positions:
        ticker = ticker_map.get(pos.symbol)
        if not ticker or not ticker.last_price or ticker.last_price <= 0:
            continue

        price = float(ticker.last_price)
        pos.current_price = price
        pos.market_value = round(pos.quantity * price, 2)

        pos.pnl = round(pos.quantity * (price - pos.cost_price), 2)
        pos.pnl_ratio = round((price - pos.cost_price) / abs(pos.cost_price) * 100, 2) if pos.cost_price != 0 else 0

        if ticker.change_24h is not None or ticker.prev_close is not None:
            # 优先用 prev_close 精确计算当日价格变动
            pc = float(ticker.prev_close) if ticker.prev_close else None
            if pc and pc > 0:
                abs_change = price - pc
            else:
                chg = float(ticker.change_24h)
                if chg != -100:
                    pc = price / (1 + chg / 100)
                    abs_change = pc * chg / 100
                else:
                    pc = price
                    abs_change = 0
            pos.day_pnl = round(pos.quantity * abs_change, 2)
            if ticker.change_24h is not None:
                pos.day_pnl_ratio = round(float(ticker.change_24h), 2)
            elif pc and pc > 0:
                pos.day_pnl_ratio = round((price - pc) / pc * 100, 2)
        else:
            pos.day_pnl = 0
            pos.day_pnl_ratio = 0

    await db.flush()


async def _refresh_position_price(db: AsyncSession, position: PositionModel) -> None:
    """用最新 ticker 价格刷新持仓的 current_price / pnl / day_pnl。"""
    try:
        from sqlalchemy import select as sync_select
        from app.core.database import SyncSessionLocal
        from app.models.market_data import Ticker as TickerModel

        sdb = SyncSessionLocal()
        try:
            # 兼容 position.symbol 不带后缀但 ticker.symbol 带后缀(.SZ/.SH)的情况
            stmt = sync_select(TickerModel).where(
                (TickerModel.symbol == position.symbol) |
                (TickerModel.symbol.like(position.symbol + ".%"))
            )
            result = sdb.execute(stmt)
            ticker = result.scalar_one_or_none()
            if ticker and ticker.last_price and ticker.last_price > 0:
                price = float(ticker.last_price)
                position.current_price = price
                position.market_value = round(position.quantity * price, 2)
                position.pnl = round(position.quantity * (price - position.cost_price), 2)
                position.pnl_ratio = round((price - position.cost_price) / abs(position.cost_price) * 100, 2) if position.cost_price != 0 else 0
                # 优先用 prev_close 精确计算当日价格变动
                if ticker.change_24h is not None or ticker.prev_close is not None:
                    pc = float(ticker.prev_close) if ticker.prev_close else None
                    if pc and pc > 0:
                        abs_change = price - pc
                    else:
                        chg = float(ticker.change_24h)
                        if chg != -100:
                            pc = price / (1 + chg / 100)
                            abs_change = pc * chg / 100
                        else:
                            pc = price
                            abs_change = 0
                    position.day_pnl = round(position.quantity * abs_change, 2)
                    if ticker.change_24h is not None:
                        position.day_pnl_ratio = round(float(ticker.change_24h), 2)
                    elif pc and pc > 0:
                        position.day_pnl_ratio = round((price - pc) / pc * 100, 2)
                else:
                    position.day_pnl = 0
                    position.day_pnl_ratio = 0
        finally:
            sdb.close()
    except Exception:
        pass


async def _fetch_positions_via_agent() -> List[Dict[str, Any]]:
    """通过 Windows easytrader 代理获取持仓。"""
    from app.utils.eastmoney_trade_adapter import EastMoneyTradeAdapter
    from app.core.config import settings

    adapter = EastMoneyTradeAdapter(agent_url=settings.EM_TRADE_AGENT_URL)
    result = await adapter.get_position()
    positions = result.get("positions", [])
    # 统一字段名：stock_code → symbol, current_amount → quantity 等
    mapped = []
    for p in positions:
        code = str(p.get("stock_code") or p.get("code") or "")
        from app.utils.eastmoney_account_client import em_code_to_symbol
        symbol = em_code_to_symbol(code)
        qty = int(float(p.get("current_amount") or p.get("enable_amount") or 0))
        if not symbol or qty <= 0:
            continue
        mapped.append({
            "symbol": symbol,
            "quantity": qty,
            "cost_price": float(p.get("cost_price") or p.get("hold_price") or 0),
            "market_value": float(p.get("market_value") or p.get("income_balance") or 0),
        })
    return mapped


async def sync_positions_from_eastmoney(
    db: AsyncSession,
    user_id: int,
) -> Dict[str, Any]:
    """从东方财富账号同步持仓到本地。

    支持两种数据源（按优先级）：
      1. EM_ACCOUNT_* cookie 直连 tradeapp.eastmoney.com（推荐，纯 Linux）
      2. Windows easytrader 代理（EM_TRADE_AGENT_URL），适合已有 PC 客户端场景

    逐条 upsert 到本地 positions 表。

    Returns:
        {created, updated, total, positions: [...]}
    """
    import logging
    from app.core.config import settings

    logger = logging.getLogger(__name__)

    raw_positions = None

    # 方式一：cookie 直连 API
    from app.utils.eastmoney_account_client import EastMoneyAccountClient
    client = EastMoneyAccountClient.from_settings()
    if client.is_configured:
        try:
            raw_positions = await client.get_positions()
        except RuntimeError as e:
            raise ValueError(f"东方财富持仓获取失败: {e}")

    # 方式二：Windows easytrader 代理
    if raw_positions is None and settings.EM_TRADE_AGENT_URL not in ("", "http://127.0.0.1:8520"):
        try:
            raw_positions = await _fetch_positions_via_agent()
        except Exception as e:
            raise ValueError(f"东方财富代理获取持仓失败: {e}")

    if raw_positions is None:
        raise ValueError(
            "东方财富账号未配置。请选择以下其一：\n"
            "  方式A: 浏览器登录 tradeapp.eastmoney.com → F12 → Application → Cookies "
            "复制 userid/ctToken/utToken/fundaccount/secuid 到 .env 的 EM_ACCOUNT_* 字段\n"
            "  方式B: Windows 端运行 eastmoney_agent.py 后，设置 EM_TRADE_AGENT_URL=http://<Windows_IP>:8520"
        )

    if not raw_positions:
        logger.info("东方财富账号无持仓数据")
        return {"created": 0, "updated": 0, "total": 0, "positions": []}

    created = 0
    updated = 0
    synced = []

    for raw in raw_positions:
        symbol = raw.get("symbol", "")
        if not symbol:
            continue

        quantity = raw.get("quantity", 0)
        cost_price = raw.get("cost_price", 0)
        if quantity <= 0:
            continue

        # 检测是否已存在
        from sqlalchemy import and_, select as sqla_select
        result = await db.execute(
            sqla_select(PositionModel).where(
                and_(
                    PositionModel.user_id == user_id,
                    PositionModel.symbol == symbol,
                )
            )
        )
        existing = result.scalar_one_or_none()

        await create_or_update_position(
            db, user_id, symbol, quantity, cost_price,
            position=existing,
        )

        if existing:
            updated += 1
        else:
            created += 1

        synced.append({
            "symbol": symbol,
            "quantity": quantity,
            "cost_price": cost_price,
            "market_value": raw.get("market_value", 0),
        })

    logger.info(
        "东方财富持仓同步完成: 新建 %s, 更新 %s, 共 %s 条",
        created, updated, len(synced),
    )
    return {
        "created": created,
        "updated": updated,
        "total": len(synced),
        "positions": synced,
    }


async def import_positions_from_excel(
    db: AsyncSession,
    user_id: int,
    file_content: bytes,
) -> dict:
    """从 Excel 文件导入持仓。返回 {total, created, updated, errors, details}。"""
    import io
    import logging
    logger = logging.getLogger(__name__)

    try:
        import openpyxl
    except ImportError:
        raise ValueError("需要安装 openpyxl 库以支持 Excel 导入")

    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    # 读取表头，识别列名
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip().lower())

    # 列名映射（支持中英文别名）
    col_map = {}
    for idx, h in enumerate(headers):
        if h in ("symbol", "代码", "标的", "股票代码", "code", "ticker"):
            col_map["symbol"] = idx
        elif h in ("quantity", "数量", "持仓", "股数", "shares", "qty"):
            col_map["quantity"] = idx
        elif h in ("cost_price", "成本价", "成本", "均价", "cost", "price"):
            col_map["cost_price"] = idx
        elif h in ("leverage", "杠杆", "倍数"):
            col_map["leverage"] = idx

    if "symbol" not in col_map:
        raise ValueError("Excel 缺少标的代码列（symbol/代码/标的）")
    if "quantity" not in col_map:
        raise ValueError("Excel 缺少数量列（quantity/数量/持仓）")
    if "cost_price" not in col_map:
        raise ValueError("Excel 缺少成本价列（cost_price/成本价/成本）")

    details = []
    created = updated = errors = 0

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c + 1).value for c in range(ws.max_column)]
        if all(v is None for v in row):
            continue  # 跳过空行

        try:
            symbol = str(row[col_map["symbol"]]).strip().upper() if row[col_map["symbol"]] is not None else ""
            if not symbol:
                details.append({"row": row_idx, "symbol": "", "status": "error", "message": "标的代码为空"})
                errors += 1
                continue

            raw_qty = row[col_map["quantity"]]
            quantity = float(raw_qty) if raw_qty is not None else 0
            if quantity <= 0:
                details.append({"row": row_idx, "symbol": symbol, "status": "error", "message": "数量必须 > 0"})
                errors += 1
                continue

            raw_price = row[col_map["cost_price"]]
            cost_price = float(raw_price) if raw_price is not None else 0
            if cost_price <= 0:
                details.append({"row": row_idx, "symbol": symbol, "status": "error", "message": "成本价必须 > 0"})
                errors += 1
                continue

            leverage = 1.0
            if "leverage" in col_map:
                raw_lev = row[col_map["leverage"]]
                if raw_lev is not None:
                    leverage = max(1.0, float(raw_lev))

            # 检查是否已存在持仓
            existing = await get_position(db, user_id, symbol)
            await create_or_update_position(db, user_id, symbol, quantity, cost_price, leverage)

            if existing:
                updated += 1
                details.append({"row": row_idx, "symbol": symbol, "status": "updated", "message": "已更新"})
            else:
                created += 1
                details.append({"row": row_idx, "symbol": symbol, "status": "created", "message": "新建成功"})

        except (ValueError, TypeError) as e:
            logger.warning("Excel 导入解析错误 第%s行: %s", row_idx, e)
            errors += 1
            details.append({"row": row_idx, "symbol": str(row[col_map.get("symbol", 0)] or ""), "status": "error", "message": str(e)})

    return {
        "total": created + updated + errors,
        "created": created,
        "updated": updated,
        "errors": errors,
        "details": details,
    }
