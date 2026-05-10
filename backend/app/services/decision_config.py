"""
决策引擎系数配置 — 从 Excel 配置文件读取，方便调参。

优先级：Excel 文件 > 下方硬编码默认值
Excel 路径：backend/config/decision_coefficients.xlsx
"""
import logging
import os
from copy import deepcopy
from typing import Any, Dict

logger = logging.getLogger(__name__)

# 配置文件路径
_CONFIG_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "config")
_CONFIG_FILE = os.path.join(_CONFIG_DIR, "decision_coefficients.xlsx")

# ── 默认系数（与 Excel 模板保持同步）─────────────────────────────────────────
# 每个系数: { "value": ..., "desc": "..." }

DEFAULT_CONFIG: Dict[str, Any] = {

    # =========================================================================
    # 1. 因子权重 — 各因子在综合评分中的权重
    # =========================================================================
    "weights": {
        "desc": "因子权重（动态权重的基础值和边界）",
        "default_technical":    {"value": 0.35, "desc": "默认：技术面权重"},
        "default_sentiment":    {"value": 0.20, "desc": "默认：情绪面权重"},
        "default_risk":         {"value": 0.25, "desc": "默认：风险面权重"},
        "default_momentum":     {"value": 0.10, "desc": "默认：动量面权重"},
        "default_fundamental":  {"value": 0.10, "desc": "默认：基本面权重"},
        # 动态权重 — 基础值
        "dyn_base_technical":   {"value": 0.35, "desc": "动态权重：技术面基础值"},
        "dyn_base_sentiment":   {"value": 0.20, "desc": "动态权重：情绪面基础值"},
        "dyn_base_risk":        {"value": 0.25, "desc": "动态权重：风险面基础值"},
        "dyn_base_momentum":    {"value": 0.10, "desc": "动态权重：动量面基础值"},
        "dyn_base_fundamental": {"value": 0.10, "desc": "动态权重：基本面基础值"},
        # 趋势强度调整
        "dyn_trend_mom_boost":      {"value": 0.18, "desc": "趋势越强，动量权重增幅（×trend_strength）"},
        "dyn_trend_fund_cut":       {"value": 0.05, "desc": "趋势越强，基本面权重降幅（×trend_strength）"},
        # 波动率调整
        "dyn_vol_risk_boost":       {"value": 0.10, "desc": "波动越高，风险权重增幅（×vol_strength）"},
        "dyn_vol_sent_boost":       {"value": 0.10, "desc": "波动越高，情绪权重增幅（×vol_strength）"},
        "dyn_vol_mom_cut":          {"value": 0.08, "desc": "波动越高，动量权重降幅（×vol_strength）"},
        # 趋势方向偏移
        "dyn_downtrend_shift_scale":{"value": 0.12, "desc": "下跌趋势：动量→情绪的偏移系数"},
        "dyn_uptrend_shift_scale":  {"value": 0.08, "desc": "上涨趋势：→动量的偏移系数"},
        "dyn_uptrend_tech_bonus":   {"value": 0.50, "desc": "上涨趋势：偏移量分配给技术面的比例"},
        # 钳位边界
        "dyn_clamp_tech_min":   {"value": 0.15, "desc": "技术面权重下限"},
        "dyn_clamp_tech_max":   {"value": 0.50, "desc": "技术面权重上限"},
        "dyn_clamp_sent_min":   {"value": 0.10, "desc": "情绪面权重下限"},
        "dyn_clamp_sent_max":   {"value": 0.40, "desc": "情绪面权重上限"},
        "dyn_clamp_risk_min":   {"value": 0.10, "desc": "风险面权重下限"},
        "dyn_clamp_risk_max":   {"value": 0.40, "desc": "风险面权重上限"},
        "dyn_clamp_mom_min":    {"value": 0.02, "desc": "动量面权重下限"},
        "dyn_clamp_mom_max":    {"value": 0.35, "desc": "动量面权重上限"},
        "dyn_clamp_fund_min":   {"value": 0.05, "desc": "基本面权重下限"},
        "dyn_clamp_fund_max":   {"value": 0.20, "desc": "基本面权重上限"},
    },

    # =========================================================================
    # 2. 趋势与波动率检测
    # =========================================================================
    "regime_detection": {
        "desc": "市场状态检测参数",
        "adx_center":           {"value": 25.0, "desc": "ADX sigmoid 中心值（>25视为有趋势）"},
        "adx_sigmoid_k":        {"value": 0.25, "desc": "ADX sigmoid 陡峭度（越小越平滑）"},
        "bb_width_center":      {"value": 0.08, "desc": "布林带宽 sigmoid 中心值（8%为中位）"},
        "bb_width_sigmoid_k":   {"value": 60.0, "desc": "布林带宽 sigmoid 陡峭度"},
        "bb_std_multiplier":    {"value": 2.0,  "desc": "布林带标准差倍数"},
        "trend_bias_scale":     {"value": 10.0, "desc": "趋势偏差放大系数（close/MA20偏离）"},
        "downtrend_threshold":  {"value": -0.02,"desc": "下跌趋势偏差阈值"},
        "uptrend_threshold":    {"value": 0.02, "desc": "上涨趋势偏差阈值"},
        "min_bars_regime":      {"value": 30,   "desc": "市场状态检测最少K线数"},
        "default_adx":          {"value": 25.0, "desc": "ADX 数据不足时的默认值"},
        "default_bb_width":     {"value": 0.05, "desc": "布林带宽数据不足时的默认值"},
        "adx_threshold_trending":{"value": 25,  "desc": "ADX>=此值视为有趋势"},
        "bb_threshold_volatile": {"value": 0.10,"desc": "布林带宽>此值视为高波动"},
        # 状态转换
        "transition_enter_mult": {"value": 1.06, "desc": "进入趋势时得分乘数（+6%）"},
        "transition_exit_mult":  {"value": 0.92, "desc": "退出趋势时得分乘数（-8%）"},
        "transition_lookback":   {"value": 5,    "desc": "状态转换检测回看K线数"},
    },

    # =========================================================================
    # 3. 综合评分与推荐阈值
    # =========================================================================
    "scoring": {
        "desc": "综合评分与推荐等级阈值",
        "baseline":             {"value": 50.0, "desc": "评分基线（0-100范围的中间值）"},
        "tanh_divisor":         {"value": 70.0, "desc": "tanh归一化除数（越小越敏感，越大越线性）"},
        "technical_daily_weight":{"value": 0.70, "desc": "日线技术评分在综合中的权重"},
        "technical_weekly_weight":{"value":0.30, "desc": "周线技术评分在综合中的权重"},
        "disagreement_penalty_scale":{"value":0.30,"desc":"信号分歧惩罚系数（×标准差）"},
        "disagreement_penalty_cap":  {"value":15.0, "desc":"信号分歧惩罚上限"},
        "agreement_factor_min": {"value": 0.40, "desc": "一致性因子下限"},
        "agreement_factor_divisor":{"value":80.0,"desc":"一致性因子计算除数（std/divisor）"},
        "confidence_floor_threshold":{"value":10, "desc":"置信度保底：复合分>此值才保底"},
        "confidence_floor_value":   {"value":10, "desc":"置信度保底值"},
        # 推荐阈值
        "rec_strong_buy":       {"value": 85,   "desc": "强烈买入阈值（>=此值）"},
        "rec_buy":              {"value": 65,   "desc": "买入阈值（>=此值）"},
        "rec_hold":             {"value": 35,   "desc": "持有阈值（>=此值，低于为卖出）"},
        "rec_sell":             {"value": 15,   "desc": "卖出阈值（<此值为强烈卖出）"},
        # 自适应阈值（信号分歧时放宽HOLD区间）
        "adaptive_hold_expansion_scale":{"value":12,"desc":"HOLD区间扩展系数（×disagreement）"},
        "adaptive_strong_barrier_scale": {"value":10,"desc":"强烈买卖阈值提升系数（×disagreement）"},
        "adaptive_buy_offset_ratio":    {"value":0.6,"desc":"买入阈值偏移比例（×hold_expansion）"},
        "adaptive_hold_offset_ratio":   {"value":0.6,"desc":"持有阈值偏移比例（×hold_expansion）"},
    },

    # =========================================================================
    # 4. 技术指标评分调整
    # =========================================================================
    "technical": {
        "desc": "技术面评分调整量（正=加分，负=减分）",
        # MA 趋势
        "ma_bullish":           {"value": 10,   "desc": "MA5>MA20 多头排列加分"},
        "ma_bearish":           {"value": -10,  "desc": "MA5<MA20 空头排列减分"},
        # MACD
        "macd_hist_up":         {"value": 8,    "desc": "MACD柱>0且上升加分"},
        "macd_hist_positive":   {"value": 4,    "desc": "MACD柱>0（稳定）加分"},
        "macd_hist_down":       {"value": -8,   "desc": "MACD柱<0且下降减分"},
        "macd_hist_negative":   {"value": -4,   "desc": "MACD柱<0（稳定）减分"},
        "macd_bull_div":        {"value": 14,   "desc": "MACD底背离（高可靠性）加分"},
        "macd_bear_div":        {"value": -14,  "desc": "MACD顶背离（高可靠性）减分"},
        "macd_min_bars":        {"value": 35,   "desc": "MACD计算最少K线数"},
        "macd_fast":            {"value": 12,   "desc": "MACD快线周期"},
        "macd_slow":            {"value": 26,   "desc": "MACD慢线周期"},
        "macd_signal":          {"value": 9,    "desc": "MACD信号线周期"},
        # RSI
        "rsi_deep_oversold":    {"value": 12,   "desc": "RSI<25深度超卖加分"},
        "rsi_oversold":         {"value": 8,    "desc": "RSI<35超卖加分"},
        "rsi_deep_overbought":  {"value": -12,  "desc": "RSI>75深度超买减分"},
        "rsi_overbought":       {"value": -8,   "desc": "RSI>65超买减分"},
        "rsi_neutral":          {"value": 2,    "desc": "RSI 35-65中性区加分"},
        "rsi_bullish_tilt":     {"value": 2,    "desc": "RSI>50偏多加分"},
        "rsi_bearish_tilt":     {"value": -2,   "desc": "RSI<=50偏空减分"},
        "rsi_period":           {"value": 14,   "desc": "RSI计算周期"},
        "rsi_min_bars":         {"value": 15,   "desc": "RSI计算最少K线数"},
        "rsi_default":          {"value": 50.0, "desc": "RSI数据不足默认值"},
        # 布林带
        "bb_lower_touch":       {"value": 8,    "desc": "触及下轨（超卖）加分"},
        "bb_upper_touch":       {"value": -8,   "desc": "触及上轨（超买）减分"},
        "bb_lower_threshold":   {"value": 1.02, "desc": "下轨触及容差（价格<=下轨×此值）"},
        "bb_upper_threshold":   {"value": 0.98, "desc": "上轨触及容差（价格>=上轨×此值）"},
        # KDJ
        "kdj_golden_oversold":  {"value": 8,    "desc": "KDJ超卖区金叉加分"},
        "kdj_death_overbought": {"value": -8,   "desc": "KDJ超买区死叉减分"},
        "kdj_golden_normal":    {"value": 4,    "desc": "KDJ普通金叉加分"},
        "kdj_death_normal":     {"value": -4,   "desc": "KDJ普通死叉减分"},
        # 成交量
        "vol_bull_div":         {"value": 12,   "desc": "成交量底背离加分"},
        "vol_bear_div":         {"value": -12,  "desc": "成交量顶背离减分"},
        "vol_ratio_high":       {"value": 1.3,  "desc": "放量阈值（当前量/均量）"},
        "vol_ratio_low":        {"value": 0.7,  "desc": "缩量阈值（当前量/均量）"},
        "vol_bull_rise":        {"value": 6,    "desc": "放量上涨加分"},
        "vol_bear_rise":        {"value": -4,   "desc": "放量下跌减分"},
        "vol_shrink":           {"value": -3,   "desc": "缩量减分"},
        "vol_short_window":     {"value": 5,    "desc": "短期均量窗口"},
        "vol_long_window":      {"value": 20,   "desc": "长期均量窗口"},
        "vol_bear_price_thresh": {"value": 1.01,"desc": "顶背离价格新高阈值（>前高×此值）"},
        "vol_bear_vol_thresh":  {"value": 0.85, "desc": "顶背离成交量阈值（<前量×此值）"},
        "vol_bull_price_thresh":{"value": 0.99, "desc": "底背离价格新低阈值（<前低×此值）"},
        "vol_bull_vol_thresh":  {"value": 1.15, "desc": "底背离成交量阈值（>前量×此值）"},
        # CMF 资金流
        "cmf_strong_buy":       {"value": 10,   "desc": "CMF>0.15 强势流入加分"},
        "cmf_mild_buy":         {"value": 4,    "desc": "CMF>0.05 温和流入加分"},
        "cmf_strong_sell":      {"value": -10,  "desc": "CMF<-0.15 强势流出减分"},
        "cmf_mild_sell":        {"value": -4,   "desc": "CMF<-0.05 温和流出减分"},
        "cmf_period":           {"value": 20,   "desc": "CMF计算周期"},
        # 信号组内折扣
        "correlation_discount": {"value": 0.50, "desc": "同组非领先信号折扣系数（×0.5=减半）"},
        "correlation_group_threshold":{"value":2,"desc":"触发折扣的组内信号数阈值"},
    },

    # =========================================================================
    # 5. 情绪面评分
    # =========================================================================
    "sentiment": {
        "desc": "情绪面评分调整量",
        "change_gt_5pct":       {"value": 15,   "desc": "24h涨幅>5%加分"},
        "change_gt_2pct":       {"value": 8,    "desc": "24h涨幅>2%加分"},
        "change_gt_0":          {"value": 3,    "desc": "24h涨幅>0加分"},
        "change_gt_minus_2pct": {"value": -3,   "desc": "24h涨幅>-2%（微跌）减分"},
        "change_gt_minus_5pct": {"value": -8,   "desc": "24h涨幅>-5%减分"},
        "change_le_minus_5pct": {"value": -15,  "desc": "24h涨幅<=-5%大幅减分"},
        "imbalance_buy":        {"value": 6,    "desc": "订单不平衡>0.2（买盘主导）加分"},
        "imbalance_sell":       {"value": -6,   "desc": "订单不平衡<-0.2（卖盘主导）减分"},
        "imbalance_threshold":  {"value": 0.2,  "desc": "订单不平衡触发阈值"},
    },

    # =========================================================================
    # 6. 风险面评分
    # =========================================================================
    "risk": {
        "desc": "风险面评分调整量",
        "risk_baseline":            {"value": 80.0, "desc": "风险评分基线（高于其他因子）"},
        "position_over_max":        {"value": -20,  "desc": "仓位超限减分"},
        "position_near_max":        {"value": -8,   "desc": "仓位逼近上限（>70%max）减分"},
        "position_near_ratio":      {"value": 0.70, "desc": "仓位逼近比例阈值"},
        "pnl_lt_minus_5pct":        {"value": -12,  "desc": "浮亏<-5%减分"},
        "pnl_lt_minus_2pct":        {"value": -5,   "desc": "浮亏<-2%减分"},
        "pnl_gt_5pct":              {"value": 5,    "desc": "浮盈>5%加分"},
        "risk_event_penalty":       {"value": -10,  "desc": "每起风控事件减分"},
        "risk_event_window_hours":  {"value": 24,   "desc": "风控事件回溯窗口（小时）"},
        "daily_loss_over_max":      {"value": -25,  "desc": "日亏损超限减分"},
        "daily_loss_near_max":      {"value": -10,  "desc": "日亏损逼近上限（>50%max）减分"},
        "daily_loss_near_ratio":    {"value": 0.50, "desc": "日亏损逼近比例阈值"},
        "default_max_position_ratio":{"value":30,   "desc": "用户未设仓位上限时的默认值(%)"},
        "default_max_daily_loss":   {"value": 5,    "desc": "用户未设日亏损上限时的默认值(%)"},
        "atr_high_vol_pct":         {"value": 0.05, "desc": "ATR>5%高波动阈值"},
        "atr_med_vol_pct":          {"value": 0.03, "desc": "ATR>3%中波动阈值"},
        "atr_high_vol_penalty":     {"value": -10,  "desc": "高波动(ATR>5%)减分"},
        "atr_med_vol_penalty":      {"value": -4,   "desc": "中波动(ATR>3%)减分"},
        "atr_low_vol_bonus":        {"value": 3,    "desc": "低波动(ATR<=3%)加分"},
        "atr_period":               {"value": 10,   "desc": "ATR计算周期"},
    },

    # =========================================================================
    # 7. 动量面评分
    # =========================================================================
    "momentum": {
        "desc": "动量面评分调整量",
        "mom_5d_gt_3pct":       {"value": 12,   "desc": "5日涨>3%加分"},
        "mom_5d_gt_1pct":       {"value": 6,    "desc": "5日涨>1%加分"},
        "mom_5d_gt_minus_1pct": {"value": -2,   "desc": "5日涨>-1%（横盘）减分"},
        "mom_5d_gt_minus_3pct": {"value": -8,   "desc": "5日涨>-3%减分"},
        "mom_5d_le_minus_3pct": {"value": -12,  "desc": "5日涨<=-3%大幅减分"},
        "ma_bull_align":        {"value": 8,    "desc": "MA5>MA10>MA20多头排列加分"},
        "ma_bear_align":        {"value": -8,   "desc": "MA5<MA10<MA20空头排列减分"},
    },

    # =========================================================================
    # 8. 基本面评分
    # =========================================================================
    "fundamental": {
        "desc": "基本面评分调整量（PE/PB/ROE/增长）",
        "pe_lt_15":             {"value": 12,   "desc": "PE<15低估值加分"},
        "pe_lt_25":             {"value": 6,    "desc": "PE<25较低估值加分"},
        "pe_lt_40":             {"value": 0,    "desc": "PE<40合理估值（不调整）"},
        "pe_lt_60":             {"value": -6,   "desc": "PE<60偏高估值减分"},
        "pe_ge_60":             {"value": -12,  "desc": "PE>=60高估减分"},
        "pb_lt_1p5":            {"value": 5,    "desc": "PB<1.5破净加分"},
        "pb_gt_8":              {"value": -5,   "desc": "PB>8高市净率减分"},
        "roe_gt_20":            {"value": 10,   "desc": "ROE>20%高盈利加分"},
        "roe_gt_15":            {"value": 6,    "desc": "ROE>15%良好盈利加分"},
        "roe_gt_8":             {"value": 2,    "desc": "ROE>8%一般盈利加分"},
        "roe_lt_3":             {"value": -4,   "desc": "ROE<3%盈利弱减分"},
        "rev_growth_gt_20":     {"value": 8,    "desc": "营收增长>20%高增长加分"},
        "rev_growth_gt_10":     {"value": 5,    "desc": "营收增长>10%稳定增长加分"},
        "rev_growth_gt_0":      {"value": 2,    "desc": "营收增长>0微增加分"},
        "rev_growth_lt_minus_10":{"value":-7,   "desc": "营收增长<-10%衰退减分"},
        "profit_growth_gt_30":  {"value": 6,    "desc": "净利润增长>30%高速加分"},
        "profit_growth_gt_10":  {"value": 3,    "desc": "净利润增长>10%稳增加分"},
        "profit_growth_lt_minus_20":{"value":-8,"desc":"净利润增长<-20%恶化减分"},
    },

    # =========================================================================
    # 9. 市场环境调整
    # =========================================================================
    "market_context": {
        "desc": "大盘环境对个股评分的调整",
        "ma60_period":          {"value": 60,   "desc": "MA60计算周期"},
        "min_bars_context":     {"value": 60,   "desc": "市场环境计算最少K线数"},
        "dev_above_15pct":      {"value": -4.0, "desc": "偏离MA60>15%（过度延伸）减分"},
        "dev_above_5pct":       {"value": 2.0,  "desc": "偏离MA60>5%（温和偏多）加分"},
        "dev_below_minus_15pct":{"value": 3.0,  "desc": "偏离MA60<-15%（超跌反弹）加分"},
        "dev_below_minus_5pct": {"value": -3.0, "desc": "偏离MA60<-5%（弱势）减分"},
    },

    # =========================================================================
    # 10. 目标价与止损
    # =========================================================================
    "target_stop": {
        "desc": "目标价和止损价计算参数",
        "atr_period_ts":        {"value": 10,       "desc": "ATR计算周期"},
        "atr_fallback_pct":     {"value": 0.02,     "desc": "ATR不可用时的回退比例（2%×价格）"},
        "target_bull_atr_mult": {"value": 2.5,      "desc": "看多目标价ATR倍数"},
        "target_moderate_atr_mult":{"value":1.5,    "desc":"中性目标价ATR倍数"},
        "stop_loss_atr_mult":   {"value": 1.5,      "desc": "止损价ATR倍数"},
        "target_bull_threshold":{"value": 60,        "desc": "触发看多目标价的评分阈值"},
        "target_moderate_threshold":{"value":40,     "desc":"触发中性目标价的评分阈值"},
    },

    # =========================================================================
    # 11. 数据获取
    # =========================================================================
    "data": {
        "desc": "数据获取相关参数",
        "klines_fresh_seconds": {"value": 21600,    "desc": "K线新鲜度阈值（6小时=21600秒）"},
        "daily_klines_limit":   {"value": 200,      "desc": "获取日线K线数量"},
        "weekly_klines_limit":  {"value": 260,      "desc": "获取周线K线数量（日线resample用）"},
        "min_klines_daily":     {"value": 20,       "desc": "日线最少K线数（不足则拒绝）"},
        "min_klines_weekly":    {"value": 10,       "desc": "周线最少K线数"},
        "mock_days":            {"value": 120,       "desc": "Mock数据生成天数"},
        "mock_volatility":      {"value": 0.02,     "desc": "Mock数据波动率"},
        "decision_valid_hours": {"value": 24,       "desc": "决策有效期（小时）"},
        "summary_top_n":        {"value": 5,        "desc": "仪表盘展示Top N推荐"},
        "summary_recent_n":     {"value": 10,       "desc": "仪表盘展示最近N条决策"},
    },
}


# ── 运行时配置（从 Excel 加载后覆盖默认值）─────────────────────────────────
_config: Dict = None


def _load_excel_config() -> Dict:
    """从 Excel 文件加载配置，覆盖默认值。返回合并后的配置。"""
    if not os.path.exists(_CONFIG_FILE):
        logger.info("决策系数配置文件不存在，使用默认值（%s）", _CONFIG_FILE)
        return deepcopy(DEFAULT_CONFIG)

    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl 未安装，使用默认系数")
        return deepcopy(DEFAULT_CONFIG)

    try:
        wb = openpyxl.load_workbook(_CONFIG_FILE, data_only=True)
    except Exception as e:
        logger.warning("无法读取决策系数配置文件 %s: %s，使用默认值", _CONFIG_FILE, e)
        return deepcopy(DEFAULT_CONFIG)

    config = deepcopy(DEFAULT_CONFIG)
    updated_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 第一行是表头: 参数名 | 当前值 | 说明
        for row_idx in range(2, ws.max_row + 1):
            name = str(ws.cell(row=row_idx, column=1).value or "").strip()
            raw_val = ws.cell(row=row_idx, column=2).value
            if not name:
                continue

            # 在默认配置中查找并覆盖
            val = float(raw_val) if raw_val is not None else None
            if val is None:
                continue

            for category, items in config.items():
                if isinstance(items, dict) and name in items:
                    if isinstance(items[name], dict) and "value" in items[name]:
                        items[name]["value"] = val
                        updated_count += 1
                    break

    if updated_count > 0:
        logger.info("从 %s 加载了 %s 个决策系数覆盖值", _CONFIG_FILE, updated_count)
    wb.close()
    return config


def get_config() -> Dict:
    """获取当前决策系数配置（懒加载，首次调用时读 Excel）。"""
    global _config
    if _config is None:
        _config = _load_excel_config()
    return _config


def get(category: str, key: str, default=None):
    """快捷取值：get("technical", "rsi_period") → 14"""
    cfg = get_config()
    cat = cfg.get(category, {})
    if isinstance(cat, dict) and "desc" in cat:
        # 跳过顶层 desc 字段
        item = cat.get(key)
        if isinstance(item, dict) and "value" in item:
            return item["value"]
    return default


def export_default_excel(path: str = None):
    """导出默认系数到 Excel 文件（用于初始化模板）。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    if path is None:
        path = _CONFIG_FILE

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for cat_name, cat_data in DEFAULT_CONFIG.items():
        ws = wb.create_sheet(title=cat_name[:31])  # Excel sheet name max 31 chars
        ws.append(["参数名", "当前值", "说明"])
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 50

        for key, item in sorted(cat_data.items()):
            if key == "desc":
                continue
            if isinstance(item, dict) and "value" in item:
                ws.append([key, item["value"], item.get("desc", "")])
            elif isinstance(item, dict) and "desc" in item:
                # 类别描述行
                pass

    wb.save(path)
    logger.info("已导出默认决策系数到 %s", path)
    return path


def reload_config():
    """强制重新加载配置（修改 Excel 后调用）。"""
    global _config
    _config = _load_excel_config()
    return _config
